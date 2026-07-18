#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Markdown成果物 → Excel納品体裁 変換ツール（Python 3.9+ / openpyxl）

正本（Markdown）から、日本のSI納品でよく求められるExcel様式
（表紙・変更履歴シート・セクション別シート・罫線つき表・A4横印刷設定）を生成する。

使い方:
    python3 md_to_excel.py 入力.md [-o 出力.xlsx] \
        --project "案件名" --author "作成者" --doc-version "1.0" \
        [--date YYYY-MM-DD] [--approver 承認者] [--font フォント名] \
        [--single-sheet] [--no-cover]

    複数ファイル: python3 md_to_excel.py a.md b.md -o 出力ディレクトリ/

変換ルール:
    # H1        → 文書名（表紙・ヘッダーに使用）
    ## H2       → シート分割（--single-sheet 指定時は1シートに連結）
    ### H3以下  → シート内の見出し行（グレー地・太字）
    | 表 |      → 罫線つき表（1行目をヘッダー装飾）
    - リスト    → ・箇条書き（ネストは全角スペースで字下げ）
    - [ ]/[x]   → ☐ / ☑
    ```code``` → 等幅フォント・グレー地
    「変更履歴」という名の H2 セクション → 表紙直後の専用シートへ
"""
import argparse
import datetime
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# ---------------------------------------------------------------- 定数（体裁）
THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")   # 表ヘッダー: 薄青
FILL_SECTION = PatternFill("solid", fgColor="E7E6E6")  # 見出し: 薄グレー
FILL_CODE = PatternFill("solid", fgColor="F2F2F2")     # コード: 薄グレー
MAX_COL_WIDTH = 60
MIN_COL_WIDTH = 6
SHEET_NAME_LIMIT = 28  # Excel上限31。連番付与の余地を残す

# ---------------------------------------------------------------- Markdown解析
RE_H = re.compile(r"^(#{1,6})\s+(.*)$")
RE_TABLE_SEP = re.compile(r"^\s*\|?\s*:?-{2,}.*$")
RE_LIST = re.compile(r"^(\s*)([-*+]|\d+[.)])\s+(.*)$")
RE_CHECK = re.compile(r"^\[([ xX])\]\s*(.*)$")


def clean_inline(text):
    """インライン記法を素のテキストに落とす（Excelセル向け）。"""
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    text = re.sub(r"(?<!\*)\*(?!\*)(.+?)(?<!\*)\*(?!\*)", r"\1", text)
    text = re.sub(r"`([^`]*)`", r"\1", text)
    text = re.sub(r"\[([^\]]+)\]\(([^)]+)\)", r"\1", text)
    text = re.sub(r"<br\s*/?>", "\n", text)
    return text.strip()


def split_table_row(line):
    row = line.strip()
    if row.startswith("|"):
        row = row[1:]
    if row.endswith("|"):
        row = row[:-1]
    return [clean_inline(c) for c in row.split("|")]


def parse_blocks(lines):
    """行のリスト → ブロック列 [(kind, payload)] に変換する。
    kind: heading(level,text) / table(rows) / list(items) / code(lines) / para(text)
    """
    blocks = []
    i = 0
    n = len(lines)
    while i < n:
        line = lines[i]
        stripped = line.strip()
        if not stripped:
            i += 1
            continue
        if stripped.startswith("```"):
            code = []
            i += 1
            while i < n and not lines[i].strip().startswith("```"):
                code.append(lines[i].rstrip())
                i += 1
            i += 1  # 閉じfence
            blocks.append(("code", code))
            continue
        m = RE_H.match(stripped)
        if m:
            blocks.append(("heading", (len(m.group(1)), clean_inline(m.group(2)))))
            i += 1
            continue
        if stripped.startswith("|") and i + 1 < n and RE_TABLE_SEP.match(lines[i + 1].strip()):
            rows = [split_table_row(stripped)]
            i += 2  # ヘッダー＋区切り行
            while i < n and lines[i].strip().startswith("|"):
                rows.append(split_table_row(lines[i].strip()))
                i += 1
            blocks.append(("table", rows))
            continue
        m = RE_LIST.match(line)
        if m:
            items = []
            while i < n:
                m2 = RE_LIST.match(lines[i])
                if not m2:
                    break
                indent = len(m2.group(1)) // 2
                text = m2.group(3)
                cm = RE_CHECK.match(text)
                if cm:
                    mark = "☑" if cm.group(1).lower() == "x" else "☐"
                    text = "%s %s" % (mark, cm.group(2))
                    bullet = ""
                elif m2.group(2)[0].isdigit():
                    bullet = m2.group(2) + " "
                else:
                    bullet = "・"
                items.append("　" * indent + bullet + clean_inline(text))
                i += 1
            blocks.append(("list", items))
            continue
        # 段落（連続行を連結）
        para = [stripped]
        i += 1
        while i < n and lines[i].strip() and not RE_H.match(lines[i].strip()) \
                and not lines[i].strip().startswith(("|", "```")) and not RE_LIST.match(lines[i]):
            para.append(lines[i].strip())
            i += 1
        blocks.append(("para", clean_inline(" ".join(para))))
    return blocks


def split_sections(blocks):
    """H2でセクション分割。先頭H1は文書名として返す。"""
    title = None
    sections = []  # (section_name, blocks)
    current_name = "概要"
    current = []
    for kind, payload in blocks:
        if kind == "heading" and payload[0] == 1 and title is None and not current and not sections:
            title = payload[1]
            continue
        if kind == "heading" and payload[0] == 2:
            if current:
                sections.append((current_name, current))
            current_name = payload[1]
            current = []
            continue
        current.append((kind, payload))
    if current:
        sections.append((current_name, current))
    return title, sections


# ---------------------------------------------------------------- Excel出力
def disp_width(text):
    """全角=2・半角=1 の表示幅。"""
    w = 0
    for ch in str(text):
        w += 2 if unicodedata.east_asian_width(ch) in ("F", "W", "A") else 1
    return w


def sanitize_sheet_name(name, used):
    name = re.sub(r"[\[\]:*?/\\]", "_", name).strip() or "Sheet"
    name = name[:SHEET_NAME_LIMIT]
    base, k = name, 2
    while name in used:
        name = "%s_%d" % (base[:SHEET_NAME_LIMIT - 3], k)
        k += 1
    used.add(name)
    return name


class SheetWriter:
    def __init__(self, ws, font_name, doc_title):
        self.ws = ws
        self.font = font_name
        self.row = 1
        self.max_col = 1
        self.col_widths = {}
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.6)
        ws.oddHeader.right.text = doc_title
        ws.oddFooter.center.text = "&P / &N"

    def _touch_width(self, col, text):
        longest = max((disp_width(s) for s in str(text).split("\n")), default=0)
        w = min(max(longest + 2, MIN_COL_WIDTH), MAX_COL_WIDTH)
        if w > self.col_widths.get(col, 0):
            self.col_widths[col] = w

    def _cell(self, col, value, **style):
        c = self.ws.cell(row=self.row, column=col, value=value)
        c.font = Font(name=self.font, size=10, bold=style.get("bold", False))
        c.alignment = Alignment(wrap_text=True, vertical="top",
                                horizontal=style.get("align", "left"))
        if style.get("fill"):
            c.fill = style["fill"]
        if style.get("border"):
            c.border = BORDER_ALL
        if style.get("mono"):
            c.font = Font(name="Courier New", size=10)
        self._touch_width(col, value or "")
        self.max_col = max(self.max_col, col)
        return c

    def heading(self, level, text):
        c = self._cell(1, text, bold=True, fill=FILL_SECTION if level <= 3 else None)
        c.font = Font(name=self.font, size=12 if level <= 3 else 10, bold=True)
        self.row += 1
        self.blank()

    def para(self, text):
        self._cell(1, text)
        self.row += 1

    def list_items(self, items):
        for it in items:
            self._cell(1, it)
            self.row += 1

    def table(self, rows):
        if not rows:
            return
        ncol = max(len(r) for r in rows)
        for ri, r in enumerate(rows):
            for ci in range(ncol):
                val = r[ci] if ci < len(r) else ""
                self._cell(ci + 1, val, bold=(ri == 0), border=True,
                           fill=FILL_HEADER if ri == 0 else None,
                           align="center" if ri == 0 else "left")
            self.row += 1
        self.blank()

    def code(self, lines):
        for ln in lines:
            self._cell(1, ln, mono=True, fill=FILL_CODE)
            self.row += 1
        self.blank()

    def blank(self):
        self.row += 1

    def finalize(self):
        for col, w in self.col_widths.items():
            self.ws.column_dimensions[get_column_letter(col)].width = w
        self.ws.print_area = "A1:%s%d" % (get_column_letter(self.max_col), max(self.row - 1, 1))


def write_blocks(writer, blocks):
    for kind, payload in blocks:
        if kind == "heading":
            writer.heading(payload[0], payload[1])
        elif kind == "table":
            writer.table(payload)
        elif kind == "list":
            writer.list_items(payload)
        elif kind == "code":
            writer.code(payload)
        else:
            writer.para(payload)


def build_cover(wb, meta, font_name):
    ws = wb.create_sheet("表紙", 0)
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9

    def put(row, col, value, size=11, bold=False, border=False, align="left", fill=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name=font_name, size=size, bold=bold)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if border:
            c.border = BORDER_ALL
        if fill:
            c.fill = fill
        return c

    ws.merge_cells("B4:H4")
    put(4, 2, meta["project"], size=14, align="center")
    ws.merge_cells("B6:H7")
    put(6, 2, meta["title"], size=22, bold=True, align="center")

    info = [("版数", meta["version"]), ("作成日", meta["date"]), ("作成者", meta["author"])]
    r = 10
    for label, value in info:
        put(r, 3, label, bold=True, border=True, align="center", fill=FILL_HEADER)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        put(r, 4, value, border=True)
        for col in range(5, 8):
            ws.cell(row=r, column=col).border = BORDER_ALL
        r += 1

    # 承認欄（承認/審査/作成の3枠。押印・サイン用に高さを取る）
    r += 2
    put(r, 3, "承認", bold=True, border=True, align="center", fill=FILL_HEADER)
    put(r, 4, "審査", bold=True, border=True, align="center", fill=FILL_HEADER)
    put(r, 5, "作成", bold=True, border=True, align="center", fill=FILL_HEADER)
    for col, name in ((3, meta.get("approver", "")), (4, ""), (5, meta["author"])):
        put(r + 1, col, name, border=True, align="center")
    ws.row_dimensions[r + 1].height = 60
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16
    return ws


def convert_file(md_path, out_path, args):
    lines = Path(md_path).read_text(encoding="utf-8").splitlines()
    title, sections = split_sections(parse_blocks(lines))
    title = args.title or title or Path(md_path).stem

    meta = {
        "title": title,
        "project": args.project,
        "version": args.doc_version,
        "date": args.date,
        "author": args.author,
        "approver": args.approver,
    }
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    if not args.no_cover:
        used_names.add("表紙")
        build_cover(wb, meta, args.font)

    # 変更履歴セクションは表紙直後の専用シートへ
    history = [(n, b) for n, b in sections if n.replace(" ", "") == "変更履歴"]
    sections = [(n, b) for n, b in sections if n.replace(" ", "") != "変更履歴"]
    if history:
        ws = wb.create_sheet(sanitize_sheet_name("変更履歴", used_names))
        w = SheetWriter(ws, args.font, title)
        for _, blocks in history:
            write_blocks(w, blocks)
        w.finalize()

    if args.single_sheet:
        ws = wb.create_sheet(sanitize_sheet_name("本文", used_names))
        w = SheetWriter(ws, args.font, title)
        for name, blocks in sections:
            if name != "概要" or blocks:
                w.heading(2, name)
                write_blocks(w, blocks)
        w.finalize()
    else:
        for name, blocks in sections:
            if name == "概要" and not blocks:
                continue
            ws = wb.create_sheet(sanitize_sheet_name(name, used_names))
            w = SheetWriter(ws, args.font, title)
            w.heading(2, name)
            write_blocks(w, blocks)
            w.finalize()

    if not wb.sheetnames:
        wb.create_sheet("本文")
    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser(description="Markdown成果物をExcel納品体裁に変換する")
    ap.add_argument("inputs", nargs="+", help="入力Markdownファイル（複数可）")
    ap.add_argument("-o", "--output", help="出力先（単一入力: .xlsxパス / 複数入力: ディレクトリ）")
    ap.add_argument("--project", default="", help="案件名（表紙）")
    ap.add_argument("--author", default="", help="作成者（表紙）")
    ap.add_argument("--approver", default="", help="承認者（表紙の承認欄）")
    ap.add_argument("--doc-version", default="1.0", help="版数（既定: 1.0）")
    ap.add_argument("--date", default=str(datetime.date.today()), help="作成日（既定: 今日）")
    ap.add_argument("--title", help="文書名（省略時はMarkdownのH1）")
    ap.add_argument("--font", default="游ゴシック", help="本文フォント（既定: 游ゴシック）")
    ap.add_argument("--single-sheet", action="store_true", help="H2でシート分割せず1シートに連結")
    ap.add_argument("--no-cover", action="store_true", help="表紙シートを作らない")
    args = ap.parse_args()

    inputs = [Path(p) for p in args.inputs]
    for p in inputs:
        if not p.is_file():
            sys.exit("入力が見つかりません: %s" % p)

    if len(inputs) == 1:
        out = Path(args.output) if args.output else inputs[0].with_suffix(".xlsx")
        if args.output and (out.is_dir() or str(args.output).endswith("/")):
            out.mkdir(parents=True, exist_ok=True)
            out = out / inputs[0].with_suffix(".xlsx").name
        print("変換: %s -> %s" % (inputs[0], convert_file(inputs[0], out, args)))
    else:
        out_dir = Path(args.output) if args.output else Path(".")
        out_dir.mkdir(parents=True, exist_ok=True)
        for p in inputs:
            out = out_dir / p.with_suffix(".xlsx").name
            print("変換: %s -> %s" % (p, convert_file(p, out, args)))


if __name__ == "__main__":
    main()
