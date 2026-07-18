#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel → Markdown 逆変換ツール（Python 3.9+ / openpyxl）

顧客から戻ってきたExcel（レビュー指摘表・修正済み設計書・課題管理表など）を
Markdownに落とし、正本（docs/配下）との差分確認・取り込みに使う。

使い方:
    python3 excel_to_md.py 入力.xlsx [-o 出力.md] [--sheets シート名 ...]

変換ルール:
    シート             → ## シート名
    連続する複数列の行 → Markdownの表（最初の行をヘッダー扱い）
    単独セルの行       → 段落
    結合セル           → 左上セルの値のみ（他は空欄）
    「表紙」シート     → 既定でスキップ（--sheets で明示すれば出力）
"""
import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

SKIP_DEFAULT = {"表紙"}


def cell_text(v):
    if v is None:
        return ""
    s = str(v).replace("\r\n", "\n").replace("\r", "\n")
    return s.replace("\n", "<br>").replace("|", "\\|").strip()


def sheet_rows(ws):
    """空行を保ったまま、各行を文字列リストにする（末尾の空セルは落とす）。"""
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = [cell_text(v) for v in row]
        while cells and not cells[-1]:
            cells.pop()
        rows.append(cells)
    while rows and not rows[-1]:
        rows.pop()
    return rows


def flush_table(buf, out):
    if not buf:
        return
    ncol = max(len(r) for r in buf)
    norm = [r + [""] * (ncol - len(r)) for r in buf]
    out.append("| " + " | ".join(norm[0]) + " |")
    out.append("|" + "---|" * ncol)
    for r in norm[1:]:
        out.append("| " + " | ".join(r) + " |")
    out.append("")
    buf.clear()


def convert_sheet(ws):
    out = ["## %s" % ws.title, ""]
    table_buf = []
    first_para = True
    for cells in sheet_rows(ws):
        nonempty = [c for c in cells if c]
        if len(nonempty) >= 2:
            table_buf.append(cells)
        else:
            flush_table(table_buf, out)
            if nonempty:
                # md_to_excel がシート先頭に書く見出しセル（=シート名）は重複するので落とす
                if first_para and nonempty[0] == ws.title:
                    first_para = False
                    continue
                first_para = False
                out.append(nonempty[0])
                out.append("")
    flush_table(table_buf, out)
    return out


def main():
    ap = argparse.ArgumentParser(description="ExcelをMarkdownに逆変換する")
    ap.add_argument("input", help="入力Excelファイル")
    ap.add_argument("-o", "--output", help="出力Markdownパス（省略時は同名.md）")
    ap.add_argument("--sheets", nargs="*", help="対象シート名（省略時は表紙以外の全シート）")
    args = ap.parse_args()

    src = Path(args.input)
    if not src.is_file():
        sys.exit("入力が見つかりません: %s" % src)
    wb = load_workbook(src, data_only=True)

    targets = args.sheets or [n for n in wb.sheetnames if n not in SKIP_DEFAULT]
    unknown = [n for n in targets if n not in wb.sheetnames]
    if unknown:
        sys.exit("シートが見つかりません: %s（存在: %s）" % (unknown, wb.sheetnames))

    lines = ["# %s" % src.stem, ""]
    for name in targets:
        lines.extend(convert_sheet(wb[name]))

    out = Path(args.output) if args.output else src.with_suffix(".md")
    out.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print("変換: %s -> %s（シート: %s）" % (src, out, ", ".join(targets)))


if __name__ == "__main__":
    main()
